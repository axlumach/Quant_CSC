import streamlit as st
import pandas as pd
import psycopg2
from psycopg2 import Error
import numpy as np
import openpyxl
import xlrd
import re
from io import BytesIO


def table_exists(cursor, table_name):

    table_name_sem_schema = table_name.split('.')[-1]

    # Query para verificar se a tabela já existe no banco de dados
    query = f"SELECT EXISTS (SELECT 1 FROM information_schema.tables WHERE table_name = '{table_name_sem_schema}');"
    
    # Executar a query
    cursor.execute(query)
    
    # Obter o resultado da consulta
    exists = cursor.fetchone()[0]
    
    return exists


def get_column_conversion_dict(sheet_name):
    conversion_dict = {}

    # Verifica o nome da planilha e define as conversões específicas
    if sheet_name == "Acoes":
        # Exemplo de conversão para a Planilha1
        conversion_dict = {
                            'data_posicao':'datetime64[ns]',
                            'carteira':str,
                            'codigo':str,
                            'papel':str,
                            'qtde_disponivel':float,
                            'qtde_bloqueada':float,
                            'qtde_total':float,
                            'custo_medio_corretagem':float,
                            'cotacao':float,
                            'cotacao_custo_medio':float,
                            'custo_total':float,
                            'resultado':float,
                            'valor_mercado':float,
                            'primeiro_venc_doador':'datetime64[ns]',
                            'qtd_tomada':float,
                            'qtd_doada':float,
                            'qtd_propria':float,
                            'perc_classe':float,
                            'perc_total':float
        }
    elif sheet_name ==  "Emprestimo_De_Acoes":
        # Exemplo de conversão para a Planilha2
        conversion_dict = {
                            'data_posicao':'datetime64[ns]',
                            'carteira':str,
                            'cliente':str,
                            'tipo_operacao':str,
                            'codigo_tipo_operacao':str,
                            'periodicidade':str,
                            'liquidacao_antecipada_doador':str,
                            'quantidade_dias_liquidacao':float,
                            'operacao_sac':float,
                            'contrato_btc':str,
                            'data_operacao':'datetime64[ns]',
                            'data_vencimento':'datetime64[ns]',
                            'data_final_carencia':'datetime64[ns]',
                            'corretora':str,
                            'papel':str,
                            'quantidade':float,
                            'preco_unitario':float,
                            'valor_financeiro':float,
                            'perc_rem':float,
                            'valor_mercado':float,
                            'remuneracao':float,
                            'apropriacao_remuneracao':float,
                            'apropriacao_taxas':float,
                            'apropriacao_total':float,
                            'modalidae_boleta':str,
                            'valor_de_comissao_fixo':float,
                            'taxa_comissao':float

        }
    elif sheet_name == "Opcoes":
        # Exemplo de conversão para a Planilha3
        conversion_dict = {
                            'data_posicao':'datetime64[ns]',
                            'carteira':str,
                            'tipo_estoque':str,
                            'codigo':str,
                            'papel':str,
                            'tipo':str,
                            'corretora':str,
                            'praca':str,
                            'exercicio':float,
                            'data_vencimento':'datetime64[ns]',
                            'quantidade_total':float,
                            'custo_medio_corretagem':float,
                            'cotacao':float,
                            'custo_total':float,
                            'resultado':float,
                            'valor_mercado':float,
                            'perc_classe':float,
                            'perc_total':float

        }
    elif sheet_name == "Opcoes_Flexiveis":
        # Exemplo de conversão para a Planilha3
        conversion_dict = {
                            'data_posicao':'datetime64[ns]',
                            'carteira':str,
                            'tipo_estoque':str,
                            'codigo':str,
                            'objeto_da_opcao':str,
                            'tipo':str,
                            'corretora':str,
                            'exercicio':float,
                            'data_vencimento':'datetime64[ns]',
                            'quantidade_total':float,
                            'custo_medio_corretagem':float,
                            'cotacao':float,
                            'custo_total':float,
                            'resultado':float,
                            'valor_mercado':float,
                            'perc_classe':float,
                            'perc_total':float

        }
    elif sheet_name == "Futuros":
        # Exemplo de conversão para a Planilha3
        conversion_dict = {
                            'data_posicao':'datetime64[ns]',
                            'carteira':str,
                            'tipo_estoque':str,
                            'ativo':str,
                            'vencimento':str,
                            'corretora':str,
                            'quantidade':float,
                            'ajuste_equalizacao':float,
                            'ajuste_valorizacao':float,
                            'preco_do_mercado':float,
                            'valor_de_mercado':float,
                            'perc_classe':float,
                            'perc_total':float

        }
    elif sheet_name == "Renda_Fixa":
        # Exemplo de conversão para a Planilha3
        conversion_dict = {
                            'data_posicao':'datetime64[ns]',
                            'carteira':str,
                            'tipo_estoque':str,
                            'codigo':str,
                            'nome':str,
                            'aplicacao':'datetime64[ns]',
                            'emitente':str,
                            'papel':float,
                            'tx_mtm':float,
                            'tx_aa':float,
                            'index':str,
                            'vencimento':'datetime64[ns]',
                            'quantidade':float,
                            'pu_de_mercado':float,
                            'valor_aplicacao':float,
                            'valor_bruto':float,
                            'impostos':float,
                            'valor_liquido':float,
                            'perc_classe':float,
                            'perc_total':float,
                            'tipo_operacao':str

        }
    elif sheet_name == "Outros_Fundos_De_Investimento":
        # Exemplo de conversão para a Planilha3
        conversion_dict = {
                            'data_posicao':'datetime64[ns]',
                            'carteira':str,
                            'tipo_estoque':str,
                            'segmento':str,
                            'codigo':str,
                            'fundo':str,
                            'instituicao':str,
                            'quantidade_cotas':float,
                            'valor_cota':float,
                            'valor_aplic_resg':float,
                            'valor_atual':float,
                            'impostos':float,
                            'valor_total_liquido':float,
                            'perc_classe':float,
                            'perc_total':float

        }
    elif sheet_name == "Conta_Corrente":
        # Exemplo de conversão para a Planilha3
        conversion_dict = {
                            'data_posicao':'datetime64[ns]',
                            'carteira':str,
                            'tipo_estoque':str,
                            'codigo':str,
                            'nm':str,
                            'instituicao':str,
                            'valor':float,
                            'perc_classe':float,
                            'perc_total':float
        }
    elif sheet_name == "Contas_Pagar_Receber":
        # Exemplo de conversão para a Planilha3
        conversion_dict = {
                            'data_posicao':'datetime64[ns]',
                            'carteira':str,
                            'tipo':float,
                            'descricao':str,
                            'valor':float,
                            'perc_classe':float,
                            'perc_total':float
        }
    elif sheet_name == "Tesouraria":
        # Exemplo de conversão para a Planilha3
        conversion_dict = {
                            'data_posicao':'datetime64[ns]',
                            'carteira':str,
                            'descricao':str,
                            'valor':float,
                            'perc_classe':float,
                            'perc_total':float
        }
    elif sheet_name == "Patrimonio_Cotas":
        # Exemplo de conversão para a Planilha3
        conversion_dict = {
                            'codigo_da_carteira':str,
                            'carteira':str,
                            'data_atual':'datetime64[ns]',
                            'atualizacao_carteira':'datetime64[ns]',
                            'data_posicao':'datetime64[ns]',
                            'patrimonio':float,
                            'valor_da_cota_bruta_de_performace':float,
                            'quantidade_de_cotas_bruta':float,
                            'quantidade_de_cotas_liq':float,
                            'vl_cota_unitaria_bruta':float,
                            'vl_cota_unitaria_liq':float,
                            'quantidade_de_cotas':float,
                            'valor_de_cotas_unitaria':float,
                            'acoes':bool,
                            'etf_rf':bool,
                            'emprestimo_de_acoes':bool,
                            'opcoes':bool,
                            'opcoes_flexiveis':bool,
                            'futuros':bool,
                            'opcoes_futuro':bool,
                            'termo':bool,
                            'renda_fixa_cpr':bool,
                            'renda_fixa':bool,
                            'outros_ativos':bool,
                            'outros_fundos_de_investimento':bool,
                            'swap':bool,
                            'conta_corrente':bool,
                            'operacoes_cambio':bool,
                            'contas_pagar_receber':bool,
                            'tesouraria':bool,
                            'rentabilidade':bool,
                            'disclaimer':bool

        }
    # else:
        # Lidar com o caso em que o nome da planilha não é reconhecido
        # print(f"Nome da planilha não reconhecido: {sheet_name}")

    return conversion_dict



# Função para obter o tipo SQL específico para cada planilha
def get_sql_type_specific_sheet(column_name, sheet_name):
    # Exemplo de mapeamento de tipos de dados específicos para cada planilha
    # Você pode adicionar lógica personalizada para cada planilha conforme necessário
    if sheet_name == "Patrimonio_Cotas":
        type_mapping = {
                        'codigo_da_carteira':'VARCHAR',
                        'carteira':'VARCHAR',
                        'data_atual':'DATE',
                        'atualizacao_carteira':'DATE',
                        'data_posicao':'DATE',
                        'patrimonio':'FLOAT8',
                        'valor_da_cota_bruta_de_performace':'FLOAT8',
                        'quantidade_de_cotas_bruta':'FLOAT8',
                        'quantidade_de_cotas_liq':'FLOAT8',
                        'vl_cota_unitaria_bruta':'FLOAT8',
                        'vl_cota_unitaria_liq':'FLOAT8',
                        'quantidade_de_cotas':'FLOAT8',
                        'valor_de_cotas_unitaria':'FLOAT8',
                        'acoes':'BOOL',
                        'etf_rf':'BOOL',
                        'emprestimo_de_acoes':'BOOL',
                        'opcoes':'BOOL',
                        'opcoes_flexiveis':'BOOL',
                        'futuros':'BOOL',
                        'opcoes_futuro':'BOOL',
                        'termo':'BOOL',
                        'renda_fixa_cpr':'BOOL',
                        'renda_fixa':'BOOL',
                        'outros_ativos':'BOOL',
                        'outros_fundos_de_investimento':'BOOL',
                        'swap':'BOOL',
                        'conta_corrente':'BOOL',
                        'operacoes_cambio':'BOOL',
                        'contas_pagar_receber':'BOOL',
                        'tesouraria':'BOOL',
                        'rentabilidade':'BOOL',
                        'disclaimer':'BOOL'
        }
    elif sheet_name == "Tesouraria":
        type_mapping = {
                        'data_posicao':'DATE',
                        'carteira':'VARCHAR',
                        'descricao':'VARCHAR',
                        'valor':'FLOAT8',
                        'perc_classe':'FLOAT8',
                        'perc_total':'FLOAT8'
        }
    elif sheet_name == "Contas_Pagar_Receber":
        type_mapping = {
                        'data_posicao':'DATE',
                        'carteira':'VARCHAR',
                        'tipo':'FLOAT8',
                        'descricao':'VARCHAR',
                        'valor':'FLOAT8',
                        'perc_classe':'FLOAT8',
                        'perc_total':'FLOAT8'            
        } 
    elif sheet_name == "Conta_Corrente":
        type_mapping = {
                        'data_posicao':'DATE',
                        'carteira':'VARCHAR',
                        'tipo_estoque':'VARCHAR',
                        'codigo':'VARCHAR',
                        'nm':'VARCHAR',
                        'instituicao':'VARCHAR',
                        'valor':'FLOAT8',
                        'perc_classe':'FLOAT8',
                        'perc_total':'FLOAT8'            
        } 
    elif sheet_name == "Outros_Fundos_De_Investimento":
        type_mapping = {
                        'data_posicao':'DATE',
                        'carteira':'VARCHAR',
                        'tipo_estoque':'VARCHAR',
                        'segmento':'VARCHAR',
                        'codigo':'VARCHAR',
                        'fundo':'VARCHAR',
                        'instituicao':'VARCHAR',
                        'quantidade_cotas':'FLOAT8',
                        'valor_cota':'FLOAT8',
                        'valor_aplic_resg':'FLOAT8',
                        'valor_atual':'FLOAT8',
                        'impostos':'FLOAT8',
                        'valor_total_liquido':'FLOAT8',
                        'perc_classe':'FLOAT8',
                        'perc_total':'FLOAT8'                              
        }   
    elif sheet_name == "Renda_Fixa":
        type_mapping = {
                        'data_posicao':'DATE',
                        'carteira':'VARCHAR',
                        'tipo_estoque':'VARCHAR',
                        'codigo':'VARCHAR',
                        'nome':'VARCHAR',
                        'aplicacao':'DATE',
                        'emitente':'VARCHAR',
                        'papel':'FLOAT8',
                        'tx_mtm':'FLOAT8',
                        'tx_aa':'FLOAT8',
                        'index':'VARCHAR',
                        'vencimento':'DATE',
                        'quantidade':'FLOAT8',
                        'pu_de_mercado':'FLOAT8',
                        'valor_aplicacao':'FLOAT8',
                        'valor_bruto':'FLOAT8',
                        'impostos':'FLOAT8',
                        'valor_liquido':'FLOAT8',
                        'perc_classe':'FLOAT8',
                        'perc_total':'FLOAT8',
                        'tipo_operacao':'VARCHAR'
        }    
    elif sheet_name == "Futuros":
        type_mapping = {
                        'data_posicao':'DATE',
                        'carteira':'VARCHAR',
                        'tipo_estoque':'VARCHAR',
                        'ativo':'VARCHAR',
                        'vencimento':'VARCHAR',
                        'corretora':'VARCHAR',
                        'quantidade':'FLOAT8',
                        'ajuste_equalizacao':'FLOAT8',
                        'ajuste_valorizacao':'FLOAT8',
                        'preco_do_mercado':'FLOAT8',
                        'valor_de_mercado':'FLOAT8',
                        'perc_classe':'FLOAT8',
                        'perc_total':'FLOAT8'                          
        }   
    elif sheet_name == "Opcoes_Flexiveis":
        type_mapping = {
                        'data_posicao':'DATE',
                        'carteira':'VARCHAR',
                        'tipo_estoque':'VARCHAR',
                        'codigo':'VARCHAR',
                        'objeto_da_opcao':'VARCHAR',
                        'tipo':'VARCHAR',
                        'corretora':'VARCHAR',
                        'exercicio':'FLOAT8',
                        'data_vencimento':'DATE',
                        'quantidade_total':'FLOAT8',
                        'custo_medio_corretagem':'FLOAT8',
                        'cotacao':'FLOAT8',
                        'custo_total':'FLOAT8',
                        'resultado':'FLOAT8',
                        'valor_mercado':'FLOAT8',
                        'perc_classe':'FLOAT8',
                        'perc_total':'FLOAT8'
        }  
    elif sheet_name == "Opcoes":
        type_mapping = {
                        'data_posicao':'DATE',
                        'carteira':'VARCHAR',
                        'tipo_estoque':'VARCHAR',
                        'codigo':'VARCHAR',
                        'papel':'VARCHAR',
                        'tipo':'VARCHAR',
                        'corretora':'VARCHAR',
                        'praca':'VARCHAR',
                        'exercicio':'FLOAT8',
                        'data_vencimento':'DATE',
                        'quantidade_total':'FLOAT8',
                        'custo_medio_corretagem':'FLOAT8',
                        'cotacao':'FLOAT8',
                        'custo_total':'FLOAT8',
                        'resultado':'FLOAT8',
                        'valor_mercado':'FLOAT8',
                        'perc_s_rv':'FLOAT8',
                        'perc_s_total':'FLOAT8'
        }   
    elif sheet_name == "Emprestimo_De_Acoes":
        type_mapping = {
                        'data_posicao':'DATE',
                        'carteira':'VARCHAR',
                        'cliente':'VARCHAR',
                        'tipo_operacao':'VARCHAR',
                        'codigo_tipo_operacao':'VARCHAR',
                        'periodicidade':'VARCHAR',
                        'liquidacao_antecipada_doador':'VARCHAR',
                        'quantidade_dias_liquidacao':'FLOAT8',
                        'operacao_sac':'FLOAT8',
                        'contrato_btc':'VARCHAR',
                        'data_operacao':'DATE',
                        'data_vencimento':'DATE',
                        'data_final_carencia':'DATE',
                        'corretora':'VARCHAR',
                        'papel':'VARCHAR',
                        'quantidade':'FLOAT8',
                        'preco_unitario':'FLOAT8',
                        'valor_financeiro':'FLOAT8',
                        'perc_rem':'FLOAT8',
                        'valor_mercado':'FLOAT8',
                        'remuneracao':'FLOAT8',
                        'apropriacao_remuneracao':'FLOAT8',
                        'apropriacao_taxas':'FLOAT8',
                        'apropriacao_total':'FLOAT8',
                        'modalidae_boleta':'VARCHAR',
                        'valor_de_comissao_fixo':'FLOAT8',
                        'taxa_comissao':'FLOAT8'
        } 
    elif sheet_name == "Acoes":
        type_mapping = {
                        'data_posicao':'DATE',
                        'carteira':'VARCHAR',
                        'codigo':'VARCHAR',
                        'papel':'VARCHAR',
                        'qtde_disponivel':'FLOAT8',
                        'qtde_bloqueada':'FLOAT8',
                        'qtde_total':'FLOAT8',
                        'custo_medio_corretagem':'FLOAT8',
                        'cotacao':'FLOAT8',
                        'cotacao_custo_medio':'FLOAT8',
                        'custo_total':'FLOAT8',
                        'resultado':'FLOAT8',
                        'valor_mercado':'FLOAT8',
                        'primeiro_venc_doador':'DATE',
                        'qtd_tomada':'FLOAT8',
                        'qtd_doada':'FLOAT8',
                        'qtd_propria':'FLOAT8',
                        'perc_classe':'FLOAT8',
                        'perc_total':'FLOAT8'
        }           
    # else:
    #     # Se a planilha não estiver especificada, use o tipo padrão
    #     type_mapping = {
    #         # Mapeamento de tipos de dados padrão...
    #     }

    # Se a coluna estiver no mapeamento, retorne o tipo de dados correspondente
    # Caso contrário, retorne o tipo de dados padrão (por exemplo, VARCHAR)
    return type_mapping.get(column_name, 'VARCHAR')




def check_and_delete_existing_data(cursor, table_name, date_column, date_value, carteira_column, carteira_value):
    # Query para verificar se existem dados com a mesma data e carteira na tabela
    query = f"SELECT EXISTS (SELECT 1 FROM {table_name} WHERE {date_column} = %s AND {carteira_column} = %s);"
    
    # Executar a query
    cursor.execute(query, (date_value, carteira_value))
    
    # Obter o resultado da consulta
    data_exists = cursor.fetchone()[0]
    
    if data_exists:
        # Se existirem dados com a mesma data e carteira, exclua-os
        delete_query = f"DELETE FROM {table_name} WHERE {date_column} = %s AND {carteira_column} = %s;"
        cursor.execute(delete_query, (date_value, carteira_value))
        print(f"Dados existentes para a data {date_value} e carteira '{carteira_value}' na tabela '{table_name}' foram excluídos.")
    else:
        print(f"Nenhum dado encontrado para a data {date_value} e carteira '{carteira_value}' na tabela '{table_name}'.")






# Função para criar a tabela a partir da planilha do Excel
def create_table_from_excel(sheet_name, excel_file, table_name, db_params, get_sql_type_func, date_as_datetimee):
    try:
        # Read the Excel sheet into a Pandas DataFrame
        df = pd.read_excel(excel_file, sheet_name=sheet_name)

        if sheet_name != "Patrimonio_Cotas":
            df.insert(0, 'data_posicao', date_as_datetimee)

        df = df.rename(columns=lambda x: x.rstrip())
        df.columns = df.columns.str.replace('Carteira/Fundo', 'carteira')
        
        df.columns = df.columns.str.replace(' ', '_').str.normalize('NFKD').str.encode('ascii', errors='ignore').str.decode('utf-8').str.lower()
        df.columns = df.columns.str.replace('.', '')
        df.columns = df.columns.str.replace('(', '')
        df.columns = df.columns.str.replace(')', '')        
        # df.columns = df.columns.map(lambda x: x.replace('.', ''))
        df.columns = df.columns.str.replace('/', '_')
        df.columns = df.columns.str.replace('ç', 'c')
        df.columns = df.columns.str.replace('%', 'perc')

        # df.columns = df.columns.str.replace('carteira_fundo', 'carteira')
        df.columns = df.columns.str.replace('perc_s_fu', 'perc_classe')
        df.columns = df.columns.str.replace('perc_s_rv', 'perc_classe')
        df.columns = df.columns.str.replace('perc_s_total', 'perc_total')
        df.columns = df.columns.str.replace('emitente_descricao', 'emitente')
        df.columns = df.columns.str.replace('tx_mtm_perc_aa','tx_mtm')
        df.columns = df.columns.str.replace('taxa_perc_aa', 'tx_aa')
        df.columns = df.columns.str.replace('perc_s_rf', 'perc_classe')
        df.columns = df.columns.str.replace('perc_s_fi', 'perc_classe')
        df.columns = df.columns.str.replace('perc_s_cpr', 'perc_classe')
        df.columns = df.columns.str.replace('perc_s_tes', 'perc_classe')
        # .replace('/', '')

        if sheet_name == "Patrimonio_Cotas":
            df['data_atual'] = pd.to_datetime(df['data_atual'], dayfirst=True)
            df['atualizacao_carteira'] = pd.to_datetime(df['atualizacao_carteira'], dayfirst=True)
            df['data_posicao'] = pd.to_datetime(df['data_posicao'], dayfirst=True)
        # if sheet_name == 'Renda_Fixa':
        #     df['aplicacao'] = pd.to_datetime(df['aplicacao'], format='%d/%m/%y')
        #     df['vencimento'] = pd.to_datetime(df['vencimento'], format='%d/%m/%y')

        if sheet_name == 'Renda_Fixa':
            df['aplicacao'] = pd.to_datetime(df['aplicacao'], dayfirst=True)
            df['vencimento'] = pd.to_datetime(df['vencimento'], dayfirst=True)
        
        if sheet_name == 'Acoes':
            df['primeiro_venc_doador'] = pd.to_datetime(df['primeiro_venc_doador'], dayfirst=True)     
            
        convert_dict = get_column_conversion_dict(sheet_name)

        # Converte as colunas de acordo com o tipo
        df = df.astype(convert_dict)

        # Establish a connection to the PostgreSQL database
        connection = psycopg2.connect(**db_params)
        
        # Create a cursor object to execute PostgreSQL commands
        cursor = connection.cursor()
        
        date_column = 'data_posicao'
        carteira_column = 'carteira'

        if table_exists(cursor, table_name):
            check_and_delete_existing_data(cursor, table_name, date_column, df[date_column].iloc[0], carteira_column, df[carteira_column].iloc[0])

        
        # Generate the CREATE TABLE SQL statement based on the DataFrame columns
        columns = ", ".join([f"{col} {get_sql_type_func(col, sheet_name)}" for col in df.columns])
        create_table_query = f"CREATE TABLE IF NOT EXISTS {table_name} ({columns});"
        
        if not table_exists(cursor, table_name):
            # Execute the CREATE TABLE SQL statement
            cursor.execute(create_table_query)
        
        # Insert data from the DataFrame into the PostgreSQL table
        for row in df.itertuples(index=False):
            row = [None if pd.isna(value) else value for value in row]
            insert_query = f"INSERT INTO {table_name} VALUES ({', '.join([f'%s' for _ in row])});"
            cursor.execute(insert_query, row)
        
        # Commit the transaction
        connection.commit()
        print(f"Table '{table_name}' created successfully in the database.")

    except (Exception, Error) as error:
        print("Error occurred while connecting to PostgreSQL:", error)

    finally:
        # Close the cursor and connection
        if connection:
            cursor.close()
            connection.close()
            print("PostgreSQL connection is closed.")





def execute_HG(excel_file_name):
        
    match = re.search(r'\d{8}', excel_file_name.name)

    date_str = match.group(0)  # Isso captura a sequência de dígitos encontrada
    
    # Formatar a string de data para o formato padrão
    formatted_date_str = f"{date_str[:4]}-{date_str[4:6]}-{date_str[6:]}"

    # Converter a string de data formatada para datetime64[ns]
    date = np.datetime64(formatted_date_str, 'ns')

    # Lista de planilhas a serem excluídas
    sheets_to_exclude = ['ETF RF', 'Opcoes_Futuro', 'Termo', 'Renda_Fixa_CPR', 'Outros_Ativos', 'SWAP', 
                          'Operacoes_cambio', 'Rentabilidade', 'Disclaimer']

    # PostgreSQL database parameters
    db_params = {
        "user": st.secrets.database.user,
        "password": st.secrets.database.password,
        "host": st.secrets.database.host,
        "port": st.secrets.database.port,
        "database": st.secrets.database.database,
    }

    if excel_file_name is not None:
        # Convertendo o UploadedFile em bytes        
        file_bytes = excel_file_name.getvalue()

    try:
    # Abra o arquivo Excel
        wb = openpyxl.load_workbook(BytesIO(file_bytes))
        print("Workbook carregado.")

    except openpyxl.utils.exceptions.InvalidFileException:
    # Se houver um InvalidFileException, tenta abrir o arquivo com xlrd (para arquivos .xls)
        try:
            wb = xlrd.open_workbook(file_contents=BytesIO(file_bytes))
            print("Passou aqui")
            
        except Exception as e:
            # Se ainda houver um erro, ele é impresso no console
            print(f"Não foi possível abrir o arquivo com xlrd devido a um erro: {e}")
    except Exception as e:
        try:
            file_bytes = excel_file_name.read()
            wb = xlrd.open_workbook(file_contents=file_bytes)            
            # wb = xlrd.open_workbook(file_contents=file_bytes.getvalue())
            print("Passou aqui")
            
        except Exception as e:
            # Se ainda houver um erro, ele é impresso no console
            print(f"Não foi possível abrir o arquivo com xlrd devido a um erro: {e}")
        
    


    # if isinstance(wb, openpyxl.workbook.workbook.Workbook):
    if isinstance(wb, openpyxl.Workbook):        
        # openpyxl para arquivos .xlsx
        print("O objeto é uma instância de Workbook.")
        sheet_names = [sheet.title for sheet in wb.worksheets if sheet.title not in sheets_to_exclude]
    # elif isinstance(wb, xlrd.book.Book):
    elif isinstance(wb, xlrd.Book):        
        # xlrd para arquivos .xls
        sheet_names = wb.sheet_names()
        sheet_names = [sheet_name for sheet_name in sheet_names if sheet_name not in sheets_to_exclude]
    else:
        print("Tipo de workbook desconhecido.")
        sheet_names = []

    # Obtenha uma lista de todos os nomes de planilhas no arquivo Excel
    # sheet_names = wb.sheetnames
    # sheet_names = [sheet_name for sheet_name in wb.sheetnames if sheet_name not in sheets_to_exclude]

    
    # Itere sobre cada nome de planilha e chame a função create_table_from_excel
    for sheet_name in sheet_names:
        # Table name in PostgreSQL
        table_name = "sch_hg." + sheet_name.lower().replace(" ", "_")  # Pode ajustar o nome da tabela conforme necessário
        
        print(date)
        print(date.dtype)
        
        # Chame a função para criar a tabela, passando a função get_sql_type_specific_sheet como argumento
        create_table_from_excel(sheet_name, excel_file_name, table_name, db_params, get_sql_type_specific_sheet, date)

    return True