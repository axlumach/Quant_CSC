import streamlit as st
import pandas as pd
import psycopg2
from psycopg2 import Error
import openpyxl
from io import BytesIO

import warnings
warnings.filterwarnings('ignore')



def read_table_to_dataframe(host, database, user, password, table_name, columns=None):
    """
    Lê uma tabela de um banco de dados PostgreSQL e retorna um DataFrame do pandas.

    Args:
    - host (str): O endereço IP do servidor onde o banco de dados está hospedado.
    - database (str): O nome do banco de dados.
    - user (str): O nome de usuário para autenticação no banco de dados.
    - password (str): A senha para autenticação no banco de dados.
    - table_name (str): O nome da tabela que será lida.
    - columns (list, opcional): Lista de nomes de colunas para selecionar. Se None, todas as colunas serão selecionadas.

    Returns:
    - DataFrame: Um DataFrame do pandas contendo os dados da tabela.
    """
    # Construir a string de conexão
    conn_string = f"host='{host}' dbname='{database}' user='{user}' password='{password}'"

    # Conectar ao banco de dados
    conn = psycopg2.connect(conn_string)

    # Criar um cursor para executar consultas
    cursor = conn.cursor()

    try:
        # Construir a consulta SQL
        if columns:
            sql_query = f"SELECT {', '.join(columns)} FROM {table_name};"
        else:
            sql_query = f"SELECT * FROM {table_name};"

        # Executar a consulta SQL
        cursor.execute(sql_query)

        # Recuperar os resultados da consulta
        data = cursor.fetchall()

        # Obter os nomes das colunas
        col_names = [desc[0] for desc in cursor.description]

        # Criar o DataFrame do pandas
        df = pd.DataFrame(data, columns=col_names)

        return df

    finally:
        # Fechar o cursor e a conexão
        cursor.close()
        conn.close()

# Exemplo de uso:
# df = read_table_to_dataframe(host='localhost', database='mydatabase', user='myuser', password='mypassword', table_name='mytable')
# print(df)




def table_exists(cursor, table_name):
    # Query para verificar se a tabela já existe no banco de dados
    query = f"SELECT EXISTS (SELECT 1 FROM information_schema.tables WHERE table_schema = 'sch_reag' AND table_name = '{table_name}');"

    print(query)
    # Executar a query
    cursor.execute(query)
    
    # Obter o resultado da consulta
    exists = cursor.fetchone()[0]
    
    return exists




def get_column_conversion_dict(sheet_name):
    conversion_dict = {}

    # Verifica o nome da planilha e define as conversões específicas
    if sheet_name == "Renda Variavel AVista":
        # Exemplo de conversão para a Planilha1
        conversion_dict = {
                            'carteira':str,
                            'data':'datetime64[ns]',
                            'tipo_posicao':str,
                            'ativo':str,
                            'descricao':str,
                            'instituicao':str,
                            'quantidade_disponivel':float,
                            'quantidade_bloqueada':float,
                            'qtd_total':float,
                            'custo_medio_s_corretagem':float,
                            'custo_medio_c_corretagem':float,
                            'custo_total_s_corretagem':float,
                            'custo_total_c_corretagem':float,
                            'resultado_c_corretagem':float,
                            'resultado_s_corretagem':float,
                            'cotacao_mercado':float,
                            'valor_bruto':float,
                            'impostos':float,
                            'valor_liquido':float,
                            'perc_patrimonio_item':float,
                            'perc_segmento_item':float
        }
    elif sheet_name ==  "Fundos Fundos":
        # Exemplo de conversão para a Planilha2
        conversion_dict = {
                            'carteira':str,
                            'data':'datetime64[ns]',
                            'tipo_posicao':str,
                            'cod':str,
                            'fundo':str,
                            'instituicao':str,
                            'quantidade_disponivel':float,
                            'quantidade_bloqueada':float,
                            'qtd_total':float,
                            'valor_cota':float,
                            'movimentos_a_converter':float,
                            'valor_bruto':float,
                            'impostos':float,
                            'valor_liquido':float,
                            'perc_patrimonio_item':float,
                            'perc_segmento_item':float
        }
    elif sheet_name == "CPR Lancamentos":
        # Exemplo de conversão para a Planilha3
        conversion_dict = {
                            'carteira':str,
                            'data':'datetime64[ns]',
                            'tipo_posicao':str,
                            'modalidade':str,
                            'cod_lancamento':str,
                            'descricao':str,
                            'valor':float,
                            'data_liquidacao':'datetime64[ns]',
                            'perc_patrimonio_item':float,
                            'perc_segmento_item':float
        }
    elif sheet_name == "Patrimonio Totais":
        # Exemplo de conversão para a Planilha3
        conversion_dict = {
                            'carteira':str,
                            'data':'datetime64[ns]',
                            'tipo_posicao':str,
                            'saldo_cpr':float,
                            'saldo_caixa_anterior':float,
                            'saldo_caixa_atual':float,
                            'valor_cota_liquido':float,
                            'valor_cota_bruta':float,
                            'quantidade_de_cotas':float,
                            'valor_patrimonio_liquido':float,
                            'valor_posicao_ativos':float,
                            'valor_cota_bruto_performance':float,
                            'valor_patrimonio_bruto_de_performance':float,
                            'valor_patrimonio_total':float,
                            'indexador_benchmark_banda':float,
                            'variacao_benchmark_banda':float,
                            'fundo_fora_de_bandas':float,
                            'status_da_banda':float,
                            'moeda':float
        }
    elif sheet_name == "Patrimonio Rentabilidade":
        # Exemplo de conversão para a Planilha3
        conversion_dict = {
                            'carteira':str,
                            'tipo_periodo':str,
                            'descricao_periodo':str,
                            'cota_liq':float 
        }
    elif sheet_name == "Renda Fixa":
        # Exemplo de conversão para a Planilha3
        conversion_dict = {
                            'carteira':str,
                            'data':'datetime64[ns]',
                            'tipo_posicao':str,
                            'tipo':str,
                            'titulo':str,
                            'cod':float,
                            'tir_operacao':float,
                            'quantidade':float,
                            'preco_unit_atual':float,
                            'valor_compra':float,
                            'valor_resgate':float,
                            'valor_bruto':float,
                            'imposto':float,
                            'valor_liquido':float,
                            'data_vencimento':'datetime64[ns]',
                            'data_vencimento_termo':'datetime64[ns]',
                            'data_compra':'datetime64[ns]',
                            'data_movimento':'datetime64[ns]',
                            'data_emissao':'datetime64[ns]',
                            'emissao':str,
                            'clearing':float,
                            'emissor':str,
                            'custodiante':str,
                            'tirmtm':float,
                            'serie':str,
                            'perc_serie':float,
                            'perc_patrimonio_item':float,
                            'perc_segmento_item':float,
                            'titulo_lastro':float,
                            'data_emissao':'datetime64[ns]',
                            'data_vencto':'datetime64[ns]'
        }
    elif sheet_name == "Caixa Lancamentos":
        # Exemplo de conversão para a Planilha3
        conversion_dict = {
                            'carteira':str,
                            'data':'datetime64[ns]',
                            'tipo_posicao':str,
                            'cod_lancamento':str,
                            'descricao':str,
                            'valor':float,
                            'conta':float,
                            'perc_patrimonio_item':float,
                            'perc_segmento_item':float
        }
    elif sheet_name == "Outros Ativos":
        # Exemplo de conversão para a Planilha3
        conversion_dict = {
                            'carteira':str,
                            'data':'datetime64[ns]',
                            'tipo_posicao':str,
                            'ativo':str,
                            'descricao':str,
                            'banco':str,
                            'valor':float,
                            'perc_patrimonio_item':float,
                            'perc_segmento_item':float

        }
    else:
        # Lidar com o caso em que o nome da planilha não é reconhecido
        print(f"Nome da planilha não reconhecido: {sheet_name}")

    return conversion_dict




# Função para obter o tipo SQL específico para cada planilha
def get_sql_type_specific_sheet(column_name, sheet_name):
    # Exemplo de mapeamento de tipos de dados específicos para cada planilha
    # Você pode adicionar lógica personalizada para cada planilha conforme necessário
    if sheet_name == "Renda Variavel AVista":
        type_mapping = {
                        'carteira':'VARCHAR',
                        'data':'DATE',
                        'tipo_posicao':'VARCHAR',
                        'ativo':'VARCHAR',
                        'descricao':'VARCHAR',
                        'instituicao':'VARCHAR',
                        'quantidade_disponivel':'float8',
                        'quantidade_bloqueada':'float8',
                        'qtd_total':'float8',
                        'custo_medio_s_corretagem':'float8',
                        'custo_medio_c_corretagem':'float8',
                        'custo_total_s_corretagem':'float8',
                        'custo_total_c_corretagem':'float8',
                        'resultado_c_corretagem':'float8',
                        'resultado_s_corretagem':'float8',
                        'cotacao_mercado':'float8',
                        'valor_bruto':'float8',
                        'impostos':'float8',
                        'valor_liquido':'float8',
                        'perc_patrimonio_item':'float8',
                        'perc_segmento_item':'float8'

            # Restante das colunas...
        }
    elif sheet_name == "Fundos Fundos":
        type_mapping = {
                        'carteira':'VARCHAR',
                        'data':'DATE',
                        'tipo_posicao':'VARCHAR',
                        'cod':'VARCHAR',
                        'fundo':'VARCHAR',
                        'instituicao':'VARCHAR',
                        'quantidade_disponivel':'float8',
                        'quantidade_bloqueada':'float8',
                        'qtd_total':'float8',
                        'valor_cota':'float8',
                        'movimentos_a_converter':'float8',
                        'valor_bruto':'float8',
                        'impostos':'float8',
                        'valor_liquido':'float8',
                        'perc_patrimonio_item':'float8',
                        'perc_segmento_item':'float8'

            # Mapeamento de tipos de dados para outra planilha...
        }
    elif sheet_name == "CPR Lancamentos":
        type_mapping = {
                        'carteira':'VARCHAR',
                        'data':'DATE',
                        'tipo_Posicao':'VARCHAR',
                        'modalidade':'VARCHAR',
                        'cod_lancamento':'VARCHAR',
                        'descricao':'VARCHAR',
                        'valor':'float8',
                        'data_liquidacao':'DATE',
                        'perc_patrimonio_item':'float8',
                        'perc_segmento_item':'float8'
            # Mapeamento de tipos de dados para outra planilha...
        } 
    elif sheet_name == "Patrimonio Totais":
        type_mapping = {
                        'carteira':'VARCHAR',
                        'data':'DATE',
                        'tipo_posicao':'VARCHAR',
                        'saldo_cpr':'float8',
                        'saldo_caixa_anterior':'float8',
                        'saldo_caixa_atual':'float8',
                        'valor_cota_liquido':'float8',
                        'valor_cota_bruta':'float8',
                        'quantidade_de_cotas':'float8',
                        'valor_patrimonio_liquido':'float8',
                        'valor_posicao_ativos':'float8',
                        'valor_cota_bruto_performance':'float8',
                        'valor_patrimonio_bruto_de_performance':'float8',
                        'valor_patrimonio_total':'float8',
                        'indexador_benchmark_banda':'float8',
                        'variacao_benchmark_banda':'float8',
                        'fundo_fora_de_bandas':'float8',
                        'status_da_banda':'float8',
                        'moeda':'float8'
            # Mapeamento de tipos de dados para outra planilha...
        } 
    elif sheet_name == "Patrimonio Rentabilidade":
        type_mapping = {
                        'carteira':'VARCHAR',
                        'tipo_periodo':'VARCHAR',
                        'descricao_periodo':'VARCHAR',
                        'cota_liq':'float8'                    
            # Mapeamento de tipos de dados para outra planilha...
        }   
    elif sheet_name == "Renda Fixa":
        type_mapping = {
                        'carteira':'VARCHAR',
                        'data':'DATE',
                        'tipo_posicao':'VARCHAR',
                        'tipo':'VARCHAR',
                        'titulo':'VARCHAR',
                        'cod':'float8',
                        'tir_operacao':'float8',
                        'quantidade':'float8',
                        'preco_unit_atual':'float8',
                        'valor_compra':'float8',
                        'valor_resgate':'float8',
                        'valor_bruto':'float8',
                        'imposto':'float8',
                        'valor_liquido':'float8',
                        'data_vencimento':'DATE',
                        'data_vencimento_termo':'DATE',
                        'data_compra':'DATE',
                        'data_movimento':'DATE',
                        'data_emissao':'DATE',
                        'emissao':'VARCHAR',
                        'clearing':'float8',
                        'emissor':'VARCHAR',
                        'custodiante':'VARCHAR',
                        'tirmtm':'float8',
                        'serie':'VARCHAR',
                        'perc_serie':'float8',
                        'perc_patrimonio_item':'float8',
                        'perc_segmento_item':'float8',
                        'titulo_lastro':'float8',
                        'data_vencto':'DATE'                   
            # Mapeamento de tipos de dados para outra planilha...
        }    
    elif sheet_name == "Caixa Lancamentos":
        type_mapping = {
                        'carteira':'VARCHAR',
                        'data':'DATE',
                        'tipo_posicao':'VARCHAR',
                        'cod_lancamento':'VARCHAR',
                        'descricao':'VARCHAR',
                        'valor':'float8',
                        'conta':'float8',
                        'perc_patrimonio_item':'float8',
                        'perc_segmento_item':'float8'                  
            # Mapeamento de tipos de dados para outra planilha...
        }   
    elif sheet_name == "Outros Ativos":
        type_mapping = {
                        'carteira':'VARCHAR',
                        'data':'DATE',
                        'tipo_posicao':'VARCHAR',
                        'ativo':'VARCHAR',
                        'descricao':'VARCHAR',
                        'banco':'VARCHAR',
                        'valor':'float8',
                        'perc_patrimonio_item':'float8',
                        'perc_segmento_item':'float8'
               
            # Mapeamento de tipos de dados para outra planilha...
        }             
    else:
        # Se a planilha não estiver especificada, use o tipo padrão
        type_mapping = {
            # Mapeamento de tipos de dados padrão...
        }

    # Se a coluna estiver no mapeamento, retorne o tipo de dados correspondente
    # Caso contrário, retorne o tipo de dados padrão (por exemplo, VARCHAR)
    return type_mapping.get(column_name, 'VARCHAR')




def check_and_delete_existing_data(cursor, table_name, date_column, date_value, carteira_column, carteira_value):
    # Query para verificar se existem dados com a mesma data e carteira na tabela
    query = f"SELECT EXISTS (SELECT 1 FROM {table_name} WHERE {date_column} = %s AND {carteira_column} = %s);"
    
    # Executar a query
    cursor.execute(query, (date_value, carteira_value))
    
    # Obter o resultado da consulta
    data_exists = cursor.fetchone()[0]
    
    if data_exists:
        # Se existirem dados com a mesma data e carteira, exclua-os
        delete_query = f"DELETE FROM {table_name} WHERE {date_column} = %s AND {carteira_column} = %s;"
        cursor.execute(delete_query, (date_value, carteira_value))
        print(f"Dados existentes para a data {date_value} e carteira '{carteira_value}' na tabela '{table_name}' foram excluídos.")
    else:
        print(f"Nenhum dado encontrado para a data {date_value} e carteira '{carteira_value}' na tabela '{table_name}'.")




# Função para criar a tabela a partir da planilha do Excel
def create_table_from_excel(sheet_name, excel_file, table_name, db_params, get_sql_type_func):
    try:
        # Read the Excel sheet into a Pandas DataFrame
        df = pd.read_excel(excel_file, sheet_name=sheet_name)

        df.columns = df.columns.str.replace(' ', '_').str.normalize('NFKD').str.encode('ascii', errors='ignore').str.decode('utf-8').str.lower()
        df.columns = df.columns.str.replace('.', '')
        df.columns = df.columns.str.replace('/', '')
        df.columns = df.columns.str.replace('ç', 'c')
        df.columns = df.columns.str.replace('%', 'perc')

        df.columns = df.columns.str.replace('quantida_de_cotas', 'quantidade_de_cotas')
        # .replace('/', '')

        # print(df.columns)

        convert_dict = get_column_conversion_dict(sheet_name)

        df = df.astype(convert_dict)

        # Converte as colunas de acordo com o tipo
        df = df.astype(convert_dict)

        # Establish a connection to the PostgreSQL database
        connection = psycopg2.connect(**db_params)
        
        # Create a cursor object to execute PostgreSQL commands
        cursor = connection.cursor()
        
        date_column = 'data'
        carteira_column = 'carteira'

        
        data_datetime = df['data'].iloc[0]
        data_formatada = data_datetime.strftime('%Y-%m-%d')
        
        table_name2 = table_name.split('.')[-1]
    
        if table_exists(cursor, table_name2):
            check_and_delete_existing_data(cursor, table_name, date_column, data_formatada, carteira_column, df['carteira'].iloc[0].split('.')[-1])            

        
        # Generate the CREATE TABLE SQL statement based on the DataFrame columns
        columns = ", ".join([f"{col} {get_sql_type_func(col, sheet_name)}" for col in df.columns])
        create_table_query = f"CREATE TABLE IF NOT EXISTS {table_name} ({columns});"
        
        if not table_exists(cursor, table_name2):
            # Execute the CREATE TABLE SQL statement
            cursor.execute(create_table_query)
        
        # Insert data from the DataFrame into the PostgreSQL table
        for row in df.itertuples(index=False):
            row = [None if pd.isna(value) else value for value in row]
            insert_query = f"INSERT INTO {table_name} VALUES ({', '.join([f'%s' for _ in row])});"
            cursor.execute(insert_query, row)
        
        # Commit the transaction
        connection.commit()
        print(f"Table '{table_name}' created successfully in the database.")

    except (Exception, Error) as error:
        print("Error occurred while connecting to PostgreSQL:", error)


def execute_REAG(excel_file_name):

    # Lista de planilhas a serem excluídas
    sheets_to_exclude = ['Patrimonio Rentabilidade']

    # PostgreSQL database parameters
    db_params = {
        "user": st.secrets.database.user,
        "password": st.secrets.database.password,
        "host": st.secrets.database.host,
        "port": st.secrets.database.port,
        "database": st.secrets.database.database,
    }

    if excel_file_name is not None:
        # Convertendo o UploadedFile em bytes
        print("Passou aqui")
        file_bytes = excel_file_name.getvalue()


    # Abra o arquivo Excel
    wb = openpyxl.load_workbook(BytesIO(file_bytes))
    print("Workbook carregado.")
    
    # Obtenha uma lista de todos os nomes de planilhas no arquivo Excel
    # sheet_names = wb.sheetnames
    sheet_names = [sheet_name for sheet_name in wb.sheetnames if sheet_name not in sheets_to_exclude]

    
    # Itere sobre cada nome de planilha e chame a função create_table_from_excel
    for sheet_name in sheet_names:
        # Table name in PostgreSQL
        table_name = "sch_reag." + sheet_name.lower().replace(" ", "_")  # Pode ajustar o nome da tabela conforme necessário
        
        
        # Chame a função para criar a tabela, passando a função get_sql_type_specific_sheet como argumento
        create_table_from_excel(sheet_name, excel_file_name, table_name, db_params, get_sql_type_specific_sheet)

    return True